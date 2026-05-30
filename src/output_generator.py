"""Output file generation for boon-academy-intervention.

Phase 4 implements write_outputs() and its private helpers for writing all output
files from the fully-enriched one-row-per-student DataFrame.

D-05 / D-06: Writes 6 output files including run_log.json built in-memory by main.py
per CONTEXT.md D-05 and D-06. The run_log dict is constructed throughout the pipeline
run and flushed once at the end via this module — no incremental writes anywhere.

D-01 (Signature): write_outputs(df, output_dir, run_log) — run_log is a required
positional parameter. Plan 04-01 implements _write_whatsapp_csv and _write_run_log.
Plans 04-02 and 04-03 add the remaining helpers and complete write_outputs().
"""
import json
import logging
import math
import re
from pathlib import Path

import pandas as pd
from fpdf import FPDF, FontFace
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src import config as cfg

logger = logging.getLogger(__name__)


def _make_pdf() -> FPDF:
    """Return a pre-configured FPDF instance: A4, 20mm margins, auto page break."""
    pdf = FPDF()
    pdf.set_margins(left=20, top=15, right=20)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    pdf.set_font("Helvetica", "", 11)
    return pdf


def _safe_text(text: str) -> str:
    """Replace non-latin1 Unicode chars with ASCII equivalents for Helvetica rendering."""
    replacements = {
        "—": " - ", "–": " - ",
        "‘": "'", "’": "'",
        "“": '"', "”": '"',
        "•": "-", "→": "->",
    }
    for char, sub in replacements.items():
        text = text.replace(char, sub)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _write_whatsapp_csv(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write whatsapp_messages.csv for all CRITICAL and HIGH risk students.

    Filters the DataFrame to CRITICAL and HIGH risk_level rows only,
    selects exactly 8 columns in the OUT-03 specified order, and writes
    a UTF-8 BOM CSV file so Excel opens it without garbled characters.

    Args:
        df: Fully enriched one-row-per-student DataFrame from enrich_with_llm().
        output_dir: Directory to write the CSV file into.

    Returns:
        Path to the written whatsapp_messages.csv file.
    """
    df_copy = df.copy()
    mask = df_copy[cfg.COL_RISK_LEVEL].isin(["CRITICAL", "HIGH"])
    cols = [
        cfg.COL_STUDENT_ID,
        cfg.COL_STUDENT_NAME,
        cfg.COL_PARENT_PHONE,
        cfg.COL_FACILITATOR_EMAIL,
        cfg.COL_CAMPUS_ID,
        cfg.COL_RISK_LEVEL,
        cfg.COL_WHATSAPP_MESSAGE,
        cfg.COL_GENERATED_BY,
    ]
    path = output_dir / "whatsapp_messages.csv"
    df_copy.loc[mask, cols].to_csv(path, index=False, encoding="utf-8-sig")
    logger.info("Wrote whatsapp CSV: %s (%d rows)", path, mask.sum())
    return path


def _write_run_log(run_log: dict, output_dir: Path) -> Path:
    """Write run_log.json from the in-memory run_log dict (D-06).

    Serializes the run_log dict to JSON with indent=2. Uses default=str
    to safely handle any datetime or Path objects that may be present
    without raising TypeError.

    Args:
        run_log: Pipeline run metadata dict with 7 required keys:
            run_timestamp, students_processed, api_calls_made, tokens_used,
            errors_encountered, fallbacks_triggered, data_quality_warnings.
        output_dir: Directory to write run_log.json into.

    Returns:
        Path to the written run_log.json file.
    """
    path = output_dir / "run_log.json"
    path.write_text(json.dumps(run_log, indent=2, default=str), encoding="utf-8")
    logger.info("Wrote run log: %s", path)
    return path


def _write_html_dashboard(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write facilitator_dashboard.html — fully self-contained HTML file (OUT-05).

    Renders the dashboard.html.j2 Jinja2 template with all student records
    embedded as a JS const (studentsData). The output file requires no server
    and no network requests — it works via file:// in any modern browser.

    JSON injection safety (CLAUDE.md critical pitfall): json.dumps().replace("</", "<\\/")
    prevents </script> in student data from breaking the HTML script tag.

    Jinja2 is loaded lazily inside this function (not at module import) to avoid
    importing jinja2 unless the dashboard helper is actually called.

    Args:
        df: Fully enriched one-row-per-student DataFrame from enrich_with_llm().
        output_dir: Directory to write facilitator_dashboard.html into.

    Returns:
        Path to the written facilitator_dashboard.html file.
    """
    from jinja2 import Environment, FileSystemLoader  # lazy import — D-01

    df_copy = df.copy()

    # Select display columns; replace NaN/NA with None so json.dumps serialises cleanly.
    # Cast to object dtype first so nullable dtypes (pd.StringDtype, pd.Int64Dtype) have
    # their pd.NA values converted to np.nan, which .where() can then replace with None
    # without the nullable array coercing None back to pd.NA (WR-01).
    display_cols = list(cfg.DISPLAY_COLS_DASHBOARD)
    display_df = df_copy[display_cols].astype(object)
    records = (
        display_df
        .where(display_df.notna(), other=None)
        .to_dict(orient="records")
    )

    # T-05-01: prevent </script> injection from student data (CLAUDE.md critical pitfall)
    students_json = json.dumps(records).replace("</", "<\\/")

    # Campus filter options — sorted unique campus IDs from data
    campus_ids: list[str] = sorted(
        df_copy[cfg.COL_CAMPUS_ID].dropna().unique().tolist()
    )

    # Load Jinja2 template from src/templates/dashboard.html.j2 (file-adjacent resource)
    template_dir = Path(__file__).parent / "templates"
    env = Environment(
        loader=FileSystemLoader(str(template_dir)),
        autoescape=False,  # students_json is pre-serialised JSON, not user HTML
    )
    template = env.get_template("dashboard.html.j2")

    run_timestamp = str(pd.Timestamp.now().isoformat())
    html = template.render(
        students_json=students_json,
        campus_ids=campus_ids,
        run_timestamp=run_timestamp,
    )

    path = output_dir / "facilitator_dashboard.html"
    path.write_text(html, encoding="utf-8")
    logger.info("Wrote HTML dashboard: %s (%d students)", path, len(df_copy))
    return path


def _str_or_na(val: object) -> str:
    """Return str(val) or 'N/A' for None / pd.NA / float NaN (WR-02).

    student.get() always returns the actual cell value when the column exists,
    so the default-parameter fallback never fires. float('nan') is truthy, so
    `or "N/A"` also fails to rescue it. This helper covers all three NA forms.
    """
    if val is None or val is pd.NA:
        return "N/A"
    if isinstance(val, float) and math.isnan(val):
        return "N/A"
    return str(val)


def _write_report(df: pd.DataFrame, run_log: dict, output_dir: Path) -> Path:
    """Write intervention_report.pdf — programmatic fpdf2 builder (OUT-04).

    Produces a 7-section A4 PDF document. Uses Helvetica (built-in) with
    _safe_text() for latin-1 safety.

    Args:
        df: Fully enriched one-row-per-student DataFrame from enrich_with_llm().
        run_log: Pipeline run metadata dict.
        output_dir: Directory to write intervention_report.pdf into.

    Returns:
        Path to the written intervention_report.pdf file.
    """
    df_copy = df.copy()
    pdf = _make_pdf()
    hs = FontFace(emphasis="BOLD", fill_color=(220, 220, 220))

    # -----------------------------------------------------------------------
    # Section 1 — Cover page
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 20)
    pdf.multi_cell(0, 10, _safe_text("Boon Academy - Student Intervention Report"), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, f"Run date: {run_log.get('run_timestamp', 'N/A')}", new_x="LMARGIN", new_y="NEXT")
    campus_count = df_copy[cfg.COL_CAMPUS_ID].nunique()
    pdf.multi_cell(0, 6, f"Campuses: {campus_count}", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(0, 6, f"Students processed: {run_log.get('students_processed', len(df_copy))}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # -----------------------------------------------------------------------
    # Section 2 — Executive Summary
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Executive Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    total = len(df_copy)
    risk_counts = df_copy[cfg.COL_RISK_LEVEL].value_counts()
    critical_count = int(risk_counts.get("CRITICAL", 0))
    high_count = int(risk_counts.get("HIGH", 0))
    medium_count = int(risk_counts.get("MEDIUM", 0))
    low_count = int(risk_counts.get("LOW", 0))
    critical_pct = (critical_count / total * 100) if total > 0 else 0.0

    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(
        0, 6,
        f"Of {total} students, {critical_count} ({critical_pct:.0f}%) are at CRITICAL "
        "risk requiring immediate intervention. The pipeline has generated prioritised "
        "facilitator action items and drafted WhatsApp parent messages for all CRITICAL "
        "and HIGH risk students.",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.ln(3)

    pdf.set_font("Helvetica", "", 10)
    with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(70, 50, 50)) as table:
        row = table.row()
        for h in ("Risk Level", "Count", "% of Total"):
            row.cell(h)
        for level, count in [
            ("CRITICAL", critical_count),
            ("HIGH", high_count),
            ("MEDIUM", medium_count),
            ("LOW", low_count),
        ]:
            pct = (count / total * 100) if total > 0 else 0.0
            row = table.row()
            row.cell(level)
            row.cell(str(count))
            row.cell(f"{pct:.1f}%")
    pdf.ln(4)

    # -----------------------------------------------------------------------
    # Section 3 — Top 10 Most At-Risk Students
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Top 10 Most At-Risk Students", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    top10 = df_copy.nlargest(10, cfg.COL_RISK_SCORE)
    pdf.set_font("Helvetica", "", 10)
    with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(15, 55, 40, 30, 30)) as table:
        row = table.row()
        for h in ("Rank", "Student Name", "Campus", "Risk Score", "Risk Level"):
            row.cell(h)
        for rank_idx, (_, student) in enumerate(top10.iterrows(), start=1):
            row = table.row()
            row.cell(str(rank_idx))
            row.cell(_safe_text(str(student[cfg.COL_STUDENT_NAME])))
            row.cell(_safe_text(str(student[cfg.COL_CAMPUS_ID])))
            row.cell(f"{student[cfg.COL_RISK_SCORE]:.1f}")
            row.cell(str(student[cfg.COL_RISK_LEVEL]))
    pdf.ln(4)

    # -----------------------------------------------------------------------
    # Section 4 — Campus Summary
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Campus Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    campus_groups = df_copy.groupby(cfg.COL_CAMPUS_ID, dropna=True)
    pdf.set_font("Helvetica", "", 10)
    with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(45, 35, 25, 25, 40)) as table:
        row = table.row()
        for h in ("Campus", "Total Students", "Critical", "High", "Intervention Coverage %"):
            row.cell(h)
        for campus_id, campus_df in campus_groups:
            c_total = len(campus_df)
            c_critical = int((campus_df[cfg.COL_RISK_LEVEL] == "CRITICAL").sum())
            c_high = int((campus_df[cfg.COL_RISK_LEVEL] == "HIGH").sum())
            c_coverage = (
                round((c_critical + c_high) / c_total * 100, 1) if c_total > 0 else 0.0
            )
            row = table.row()
            row.cell(_safe_text(str(campus_id)))
            row.cell(str(c_total))
            row.cell(str(c_critical))
            row.cell(str(c_high))
            row.cell(f"{c_coverage}%")
    pdf.ln(4)

    # -----------------------------------------------------------------------
    # Section 5 — Student Deep-Dives
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Student Deep-Dives", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    component_cols = [
        cfg.COL_ATTENDANCE_RATE,
        cfg.COL_AVG_PRACTICE,
        cfg.COL_TREND_DIR,
        cfg.COL_DAYS_SINCE_NOTE,
    ]
    component_labels = [
        "Attendance Rate",
        "Avg Practice Questions",
        "Trend Direction",
        "Days Since Last Note",
    ]

    for tier in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        tier_df = df_copy[df_copy[cfg.COL_RISK_LEVEL] == tier]
        if tier_df.empty:
            continue
        student = tier_df.nlargest(1, cfg.COL_RISK_SCORE).iloc[0]
        pdf.set_font("Helvetica", "B", 12)
        pdf.multi_cell(0, 7, _safe_text(f"{tier} - {student[cfg.COL_STUDENT_NAME]}"), new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("Helvetica", "", 11)
        pdf.multi_cell(
            0, 6,
            _safe_text(
                f"Campus: {student[cfg.COL_CAMPUS_ID]} | "
                f"Risk Score: {student[cfg.COL_RISK_SCORE]:.1f} | "
                f"Level: {student[cfg.COL_RISK_LEVEL]}"
            ),
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(100, 70)) as table:
            row = table.row()
            row.cell("Component")
            row.cell("Score")
            for col, label in zip(component_cols, component_labels):
                value = student[col] if col in df_copy.columns else "N/A"
                row = table.row()
                row.cell(label)
                row.cell(_safe_text(str(value)))
        pdf.set_font("Helvetica", "", 11)
        pdf.multi_cell(
            0, 6,
            f"Facilitator Summary: {_safe_text(_str_or_na(student[cfg.COL_FACILITATOR_SUMMARY]))}",
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.multi_cell(
            0, 6,
            f"Recommended Action: {_safe_text(_str_or_na(student[cfg.COL_RECOMMENDED_ACTION]))}",
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(3)

    # -----------------------------------------------------------------------
    # Section 6 — Automated Data Cleanup Summary
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Automated Data Cleanup Summary", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    warnings = run_log.get("data_quality_warnings", [])
    missing_count = sum(1 for w in warnings if w.get("type") == "missing_numeric")
    mismatch_count = sum(1 for w in warnings if w.get("type") == "type_mismatch")
    duplicate_count = sum(1 for w in warnings if w.get("type") == "duplicate_id")

    pdf.set_font("Helvetica", "", 11)
    if not warnings:
        pdf.multi_cell(
            0, 6,
            "No data quality issues detected. All student records were complete and valid.",
            new_x="LMARGIN", new_y="NEXT",
        )
    else:
        pdf.multi_cell(
            0, 6,
            "The pipeline automatically detected and resolved the following data quality "
            "issues before scoring. No students were excluded - all records were corrected "
            "and included in the risk analysis.",
            new_x="LMARGIN", new_y="NEXT",
        )
        pdf.ln(2)
        pdf.set_font("Helvetica", "", 10)
        with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(80, 30, 60)) as table:
            row = table.row()
            for h in ("Issue Type", "Count", "Action Taken"):
                row.cell(h)
            for issue_type, count, action in [
                ("Missing numeric values", str(missing_count), "Filled with 0"),
                ("Non-numeric values", str(mismatch_count), "Coerced to 0"),
                ("Duplicate student IDs", str(duplicate_count), "Kept last record"),
            ]:
                row = table.row()
                row.cell(issue_type)
                row.cell(count)
                row.cell(action)
    pdf.ln(4)

    # -----------------------------------------------------------------------
    # Section 7 — Methodology Appendix
    # -----------------------------------------------------------------------
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(0, 8, "Methodology Appendix", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(0, 7, "Risk Score Formula", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, "Risk score (0-100) is computed as a weighted sum of four components:", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 10)
    with pdf.table(headings_style=hs, num_heading_rows=1, col_widths=(120, 50)) as table:
        row = table.row()
        row.cell("Component")
        row.cell("Weight")
        for component, weight in [
            ("Attendance Rate", f"{cfg.WEIGHT_ATTENDANCE:.0%}"),
            ("Avg Practice Questions", f"{cfg.WEIGHT_PRACTICE:.0%}"),
            ("Trend Direction", f"{cfg.WEIGHT_TREND:.0%}"),
            ("Days Since Last Note", f"{cfg.WEIGHT_NOTES:.0%}"),
        ]:
            row = table.row()
            row.cell(component)
            row.cell(weight)
    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 12)
    pdf.multi_cell(0, 7, "Risk Level Thresholds", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(0, 6, f"CRITICAL: score >= {cfg.RISK_THRESHOLD_CRITICAL}", new_x="LMARGIN", new_y="NEXT")
    pdf.multi_cell(
        0, 6, f"HIGH: {cfg.RISK_THRESHOLD_HIGH} <= score < {cfg.RISK_THRESHOLD_CRITICAL}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.multi_cell(
        0, 6, f"MEDIUM: {cfg.RISK_THRESHOLD_MEDIUM} <= score < {cfg.RISK_THRESHOLD_HIGH}",
        new_x="LMARGIN", new_y="NEXT",
    )
    pdf.multi_cell(0, 6, f"LOW: score < {cfg.RISK_THRESHOLD_MEDIUM}", new_x="LMARGIN", new_y="NEXT")

    path = output_dir / "intervention_report.pdf"
    pdf.output(str(path))
    logger.info("Wrote PDF report: %s (%d students)", path, len(df_copy))
    return path


def _write_priority_list(df: pd.DataFrame, output_dir: Path) -> Path:
    """Write intervention_priority_list.xlsx — all students ranked by risk_score desc.

    Produces a 12-column Excel file (OUTPUT_COLS_PRIORITY) with:
    - Row 1: navy header (COLOR_HEADER fill, white bold font)
    - Rows 2+: data rows color-coded by risk_level
    - freeze_panes="A2" — header row always visible when scrolling
    - Auto column widths capped at 60 chars + 2 padding

    Args:
        df: Fully enriched one-row-per-student DataFrame.
        output_dir: Directory to write the Excel file into.

    Returns:
        Path to the written intervention_priority_list.xlsx file.
    """
    df_copy = df.copy()

    # Sort descending by risk_score and assign sequential ranks
    df_sorted = df_copy.sort_values(cfg.COL_RISK_SCORE, ascending=False).reset_index(
        drop=True
    )
    df_sorted[cfg.COL_RANK] = df_sorted.index + 1

    # Select the 12 output columns in the correct order
    df_out = df_sorted[list(cfg.OUTPUT_COLS_PRIORITY)]

    # Create workbook and write header row (row 1)
    wb = Workbook()
    ws = wb.active
    ws.title = "Intervention Priority List"
    header_fill = PatternFill(fill_type="solid", fgColor=cfg.COLOR_HEADER)
    header_font = Font(bold=True, color=cfg.FONT_WHITE)
    for col_idx, col_name in enumerate(cfg.OUTPUT_COLS_PRIORITY, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font

    # Build fill map once before the data-row loop (not per-cell — see research pitfall)
    fill_map = {
        "CRITICAL": PatternFill(fill_type="solid", fgColor=cfg.COLOR_CRITICAL),
        "HIGH": PatternFill(fill_type="solid", fgColor=cfg.COLOR_HIGH),
        "MEDIUM": PatternFill(fill_type="solid", fgColor=cfg.COLOR_MEDIUM),
        "LOW": PatternFill(fill_type="solid", fgColor=cfg.COLOR_LOW),
    }

    # Write data rows starting at row 2
    for row_idx, (_, row) in enumerate(df_out.iterrows(), start=2):
        risk_level = row[cfg.COL_RISK_LEVEL]
        row_fill = fill_map.get(risk_level)
        for col_idx, col_name in enumerate(cfg.OUTPUT_COLS_PRIORITY, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            if row_fill:
                cell.fill = row_fill

    # Auto column widths — None guard is mandatory (str(None)=="None" underestimates)
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
            min(max_len, 60) + 2
        )

    # Freeze header row and save
    ws.freeze_panes = "A2"
    path = output_dir / "intervention_priority_list.xlsx"
    wb.save(path)
    logger.info("Wrote priority list: %s (%d rows)", path, len(df_out))
    return path


def _write_campus_dashboards(
    df: pd.DataFrame, output_dir: Path
) -> dict[str, Path]:
    """Write one facilitator_dashboard_{campus_id}.xlsx per unique campus.

    Each campus file has 15 columns (OUTPUT_COLS_CAMPUS = 12 standard + 3 LLM):
    - Row 1: navy header (same styling as priority list)
    - Row 2: summary stats row (bold, light grey fill) — total, CRITICAL, HIGH, coverage%
    - Rows 3+: student data rows color-coded by risk_level
    - freeze_panes="A2" — only the header row is frozen (D-08)
    - MEDIUM/LOW rows have None in the 3 LLM columns (D-06 — empty, not "N/A")
    - NaN campus_id rows are excluded via dropna=True in groupby

    Args:
        df: Fully enriched one-row-per-student DataFrame.
        output_dir: Directory to write the Excel files into.

    Returns:
        Dict mapping "campus_{campus_id}" keys to Path objects for each written file.
    """
    df_copy = df.copy()
    results: dict[str, Path] = {}

    # LLM column names — cells for MEDIUM/LOW students are written as None (D-06)
    llm_col_names = {
        cfg.COL_FACILITATOR_SUMMARY,
        cfg.COL_WHATSAPP_MESSAGE,
        cfg.COL_GENERATED_BY,
    }

    # Build fill map once (shared across all campus files)
    fill_map = {
        "CRITICAL": PatternFill(fill_type="solid", fgColor=cfg.COLOR_CRITICAL),
        "HIGH": PatternFill(fill_type="solid", fgColor=cfg.COLOR_HIGH),
        "MEDIUM": PatternFill(fill_type="solid", fgColor=cfg.COLOR_MEDIUM),
        "LOW": PatternFill(fill_type="solid", fgColor=cfg.COLOR_LOW),
    }

    # Group by campus, excluding NaN campus_ids (dropna=True prevents _nan.xlsx files)
    for campus_id, campus_df in df_copy.groupby(cfg.COL_CAMPUS_ID, dropna=True):
        campus_df = campus_df.sort_values(
            cfg.COL_RISK_SCORE, ascending=False
        ).reset_index(drop=True)
        campus_df[cfg.COL_RANK] = campus_df.index + 1

        # Compute summary stats for row 2
        total = len(campus_df)
        critical_count = int((campus_df[cfg.COL_RISK_LEVEL] == "CRITICAL").sum())
        high_count = int((campus_df[cfg.COL_RISK_LEVEL] == "HIGH").sum())
        coverage_pct = (
            round((critical_count + high_count) / total * 100, 1) if total > 0 else 0.0
        )

        # Select 15 output columns
        df_out = campus_df[list(cfg.OUTPUT_COLS_CAMPUS)]

        # Create workbook and write header row (row 1)
        wb = Workbook()
        ws = wb.active
        ws.title = str(campus_id)[:31]  # openpyxl sheet name limit (WR-03)
        header_fill = PatternFill(fill_type="solid", fgColor=cfg.COLOR_HEADER)
        header_font = Font(bold=True, color=cfg.FONT_WHITE)
        for col_idx, col_name in enumerate(cfg.OUTPUT_COLS_CAMPUS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font

        # Write summary row (row 2) — bold, light grey fill
        summary_fill = PatternFill(fill_type="solid", fgColor="FFEEEEEE")
        summary_font = Font(bold=True)
        summary_values = [
            "Summary",
            "",
            "",
            str(campus_id),
            "",
            "",
            "",
            f"Total: {total}",
            f"CRITICAL: {critical_count}",
            f"HIGH: {high_count}",
            f"Coverage: {coverage_pct}%",
            "",
            "",
            "",
            "",
        ]
        for col_idx, val in enumerate(summary_values, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill = summary_fill
            cell.font = summary_font

        # Write data rows starting at row 3
        for row_idx, (_, row) in enumerate(df_out.iterrows(), start=3):
            risk_level = row[cfg.COL_RISK_LEVEL]
            row_fill = fill_map.get(risk_level)
            is_llm_eligible = risk_level in ("CRITICAL", "HIGH")
            for col_idx, col_name in enumerate(cfg.OUTPUT_COLS_CAMPUS, start=1):
                if col_name in llm_col_names and not is_llm_eligible:
                    value = None  # D-06: empty cell, not "N/A"
                else:
                    value = row[col_name]
                    # Normalise pandas NA / float NaN to Python None for openpyxl (CR-03)
                    if value is pd.NA or value is None:
                        value = None
                    elif isinstance(value, float) and pd.isna(value):
                        value = None  # float('nan') caught before the isinstance guard
                    elif not isinstance(value, (str, int, float, bool)):
                        try:
                            if pd.isna(value):
                                value = None
                        except (TypeError, ValueError):
                            pass
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_fill:
                    cell.fill = row_fill

        # Auto column widths (same pattern as priority list)
        for col_cells in ws.columns:
            max_len = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = (
                min(max_len, 60) + 2
            )

        # Freeze header row only (D-08: freeze_panes="A2")
        ws.freeze_panes = "A2"
        # Sanitise campus_id before embedding in filename — prevents path traversal
        # for untrusted CSV input (WR-03); re imported at module level
        safe_campus_id = re.sub(r'[^\w\-]', '_', str(campus_id))
        path = output_dir / f"facilitator_dashboard_{safe_campus_id}.xlsx"
        wb.save(path)
        logger.info("Wrote campus dashboard: %s (%d students)", path, total)
        results[f"campus_{campus_id}"] = path

    return results


def write_outputs(
    df: pd.DataFrame,
    output_dir: Path,
    run_log: dict,
) -> dict[str, Path]:
    """Write all output files (Phase 4: Excel/CSV/JSON; Phase 5: HTML dashboard and Word report).

    Orchestrates six private helpers — one per output type — and returns a unified
    dict mapping semantic keys to resolved Path objects (D-01, D-02).

    D-03: Creates output_dir (including parents) if it does not exist — idempotent.
    D-04: Delegates to six independently-testable private helpers.

    Args:
        df: Fully enriched one-row-per-student DataFrame from enrich_with_llm().
        output_dir: Directory to write all output files into (cfg.OUTPUT_DIR).
        run_log: In-memory run metadata dict built throughout the pipeline run (D-05/D-06).
                 Keys: run_timestamp, students_processed, api_calls_made, tokens_used,
                 errors_encountered, fallbacks_triggered, data_quality_warnings.

    Returns:
        Dict mapping output file keys to their resolved Path objects.
        Keys: "priority_list", "campus_{campus_id}" per campus, "whatsapp", "run_log",
              "dashboard" (Path to facilitator_dashboard.html, OUT-05),
              "report" (Path to intervention_report.docx, OUT-04).
    """
    output_dir.mkdir(parents=True, exist_ok=True)   # D-03

    if df.empty:
        logger.warning(
            "write_outputs called with empty DataFrame — no student data to write"
        )

    paths: dict[str, Path] = {}

    priority_path = _write_priority_list(df, output_dir)
    paths["priority_list"] = priority_path

    campus_paths = _write_campus_dashboards(df, output_dir)
    paths.update(campus_paths)

    whatsapp_path = _write_whatsapp_csv(df, output_dir)
    paths["whatsapp"] = whatsapp_path

    run_log_path = _write_run_log(run_log, output_dir)
    paths["run_log"] = run_log_path

    dashboard_path = _write_html_dashboard(df, output_dir)
    paths["dashboard"] = dashboard_path

    report_path = _write_report(df, run_log, output_dir)
    paths["report"] = report_path

    logger.info(
        "All outputs written to %s — keys: %s",
        output_dir,
        list(paths.keys()),
    )
    return paths
